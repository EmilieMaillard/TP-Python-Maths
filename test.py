import collections
from openpyxl import Workbook, load_workbook

##### Etape 1 : On lit et on ouvre le fichier #####

wb = load_workbook('donnees.xlsx') 

#on spécifie le nom de la feuille sur laquelle on veut travailler ou on peut utiliser la fonction .active pour spéficier la feuille active
ws = wb['data'] 

##### Etape 2 : on créer un dictionnaire pour ajouter les données du fichier (les factures sont les clés et les produits sont des listes de valeurs correspondant à chaque clé)
myDict = collections.defaultdict(list) 

for row in ws.rows:                                 #pour chacune des lignes dans la feuille concernée
    myDict[row[0].value].append(str(row[1].value))  #on renvoie le dictionnaire avec la clé puis ensuite une liste des valeurs 

##### Etape 3 : Récupération de la liste de tous les produits #####

column = ws['B']

for i in range(len(column)):
    print(column[i].value)


##### Etape 4 : Création d'un nouveau classeur et d'une nouvelle feuille 
# - on affiche les valeurs des produits sous forme d'un matrice 
# - on fait les calculs à partir du dictionnaire
# - on sauvegarde le fichier #####


#on crée un nouveau classeur dans un nouveau fichier resultat.xlsx
wb_writer = Workbook(write_only=True)

#on crée une nouvelle feuille dans ce nouveau classeur
ws_writer = wb_writer.create_sheet()

   


wb_writer.save(filename= 'resultat.xlsx') #sauvegarder le fichier 